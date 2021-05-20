from selenium import webdriver
from lxml import html, etree
import time
import sys
from PyQt5.QtWidgets import QVBoxLayout, QPlainTextEdit, QWidget, QMainWindow, QApplication, QWidget, QPushButton, QAction, QLineEdit, QMessageBox, QLabel,QTableWidget,QTableWidgetItem
from PyQt5.QtGui import QIcon, QPixmap
from PyQt5.QtCore import pyqtSlot
import xlsxwriter
from bs4 import BeautifulSoup
import urllib.request
from random import randint
import os
from selenium.webdriver.chrome.options import Options
from selenium.webdriver import ActionChains
import requests
from selenium.webdriver.common.action_chains import ActionChains
import math
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import subprocess
import pyperclip



workbook = xlsxwriter.Workbook('ali.xlsx')

class App(QMainWindow):

    def update_label(self):
        self.update = update.label_update(self)

    def __init__(self):
        super().__init__()
        self.title = 'Aliexpress parser'
        self.left = 10
        self.top = 10
        self.width = 1050
        self.height = 620
        self.initUI()
    
    def initUI(self):
        global label
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)

        # Create textbox
        self.textbox = QLineEdit(self)
        self.textbox.move(200, 40)
        self.textbox.resize(500,15)
        
        # Create a button in the window
        self.button = QPushButton('Parse', self)
        self.button.move(875, 33)
        self.button.resize(150,50)

        self.findkeys = QPushButton('Find keys', self)
        self.findkeys.move(875, 515)
        self.findkeys.resize(100,25)
        self.findkeys.clicked.connect(self.getkeys)

        self.deleteduplicates = QPushButton('Delete duplicates', self)
        self.deleteduplicates.move(875, 540)
        self.deleteduplicates.resize(130,25)

        self.toexcel = QPushButton('Load to excel', self)
        self.toexcel.move(875, 570)
        self.toexcel.resize(130,40)
        self.toexcel.clicked.connect(self.toexcelf)

        self.name = QPlainTextEdit(self)
        self.name.move(200, 70)
        self.name.resize(670, 45)

        self.description = QPlainTextEdit(self)
        self.description.move(550, 150)
        self.description.resize(320, 350)

        self.feature1 = QPlainTextEdit(self)
        self.feature1.move(200, 150)
        self.feature1.resize(280, 50)

        self.feature2 = QPlainTextEdit(self)
        self.feature2.move(200,210)
        self.feature2.resize(280, 50)

        self.feature3 = QPlainTextEdit(self)
        self.feature3.move(200, 270)
        self.feature3.resize(280, 50)

        self.feature4 = QPlainTextEdit(self)
        self.feature4.move(200, 330)
        self.feature4.resize(280, 50)

        self.feature5 = QPlainTextEdit(self)
        self.feature5.move(200, 390)
        self.feature5.resize(280, 50)

        self.keys = QPlainTextEdit(self)
        self.keys.move(200, 520)
        self.keys.resize(670, 75)

        featlabel = QLabel('Features', self)
        featlabel.move(200,120)

        label = QLabel('Description', self)
        label.move(550,120)

        alistandart = QPushButton('', self)
        alistandart.move(190,5)
        alistandart.resize(30,30)
        alistandart.clicked.connect(self.standparse)

        labelship1 = QLabel('Aliexpress Standart Shipping', self)
        labelship1.adjustSize()
        labelship1.move(220,12)

        alistandart2 = QPushButton('', self)
        alistandart2.move(440,5)
        alistandart2.resize(30,30)
        alistandart2.clicked.connect(self.standparse2)

        labelship2 = QLabel('Cainiao Super Economy Global', self)
        labelship2.adjustSize()
        labelship2.move(470,12)

        labelpercent = QLabel('Процент наценки:  ', self)
        labelpercent.adjustSize()
        labelpercent.move(690,12)

        labelprice = QLabel('Цена: ', self)
        labelprice.adjustSize()
        labelprice.move(870,12)

        self.pricebox = QLineEdit(self)
        self.pricebox.move(915, 12)
        self.pricebox.resize(100,20)

        self.percentbox = QLineEdit(self)
        self.percentbox.move(810, 12)
        self.percentbox.resize(40,20)

        # connect button to function on_click
        self.button.clicked.connect(self.on_click)
        self.show()

    @pyqtSlot()
    def on_click(self):
        global p
        global driver
        global imgs

        link = self.textbox.text()
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument('lang=en')

        driver = webdriver.Chrome(options=chrome_options)
        driver.get(link)
        
        time.sleep(4)
        
        element_to_click = driver.find_element_by_xpath("//span[@class=\"language_txt\"]")
        actions = ActionChains(driver)
        actions.click(element_to_click)
        # perform the operation on the element
        actions.perform()

        element_to_click = driver.find_element_by_xpath("//div[@class=\"switcher-currency-c language-selector\"]")
        actions = ActionChains(driver)
        actions.click(element_to_click)
	    # perform the operation on the element
        actions.perform()

        element_to_click = driver.find_element_by_xpath("//a[@data-locale=\"en_US\"]")
        actions = ActionChains(driver)
        actions.click(element_to_click)
	    # perform the operation on the element
        actions.perform()

        element_to_click = driver.find_element_by_xpath("//button[@data-role=\"save\"]")
        actions = ActionChains(driver)
        actions.click(element_to_click)
	    # perform the operation on the element
        actions.perform()

        time.sleep(1)


        driver.execute_script("window.scrollTo(0,900)")

        time.sleep(5)

        html = driver.page_source
        html = BeautifulSoup(html,features="lxml")
        details = html.find('div', {'class':'detailmodule_html'})


        details = str(details)
        details = BeautifulSoup(details, "lxml")
        details = details.get_text()

        images = html.find('div', {'class':'images-view-wrap'})
        images = images.findAll('div', {'class':'images-view-item'})
        links = []

        title = html.find('h1', {'class':"product-title-text"})
        title = str(title)
        title = BeautifulSoup(title, "lxml")
        title = title.get_text()

        if p == 0:
            self.name.insertPlainText(title)
            p += 1
        else:
            self.name.clear()
            self.name.insertPlainText(title)
            p+=1
        title = self.name.toPlainText()


        for i in range(len(images)):
            links.append(images[i].find('img')["src"])
        
        try:
            for i in range(len(images)):
                r = urllib.request.urlopen(links[i])
                n = randint(1,10000000)
                with open(f"img{i}{n}.jpg", "wb") as f:
                    f.write(r.read())
                self.label = QLabel(self)
                self.label.resize(100,100)
                self.pixmap = QPixmap(f'img{i}{n}.jpg')
                self.label.setPixmap(self.pixmap)
                self.label.move(40,50+i*60+25)
                os.remove(f'img{i}{n}.jpg')
                self.label.show()
        except Exception:
            print("")

        if p == 0:
            self.description.insertPlainText(details)
            p += 1
        else:
            self.description.clear()
            self.description.insertPlainText(details)
            p+=1
        self.description.show()

    @pyqtSlot()
    def standparse(self):
        global driver
        try:
            element_to_click = driver.find_element_by_xpath("//span[@class=\"product-shipping-info black-link\"]")
            actions = ActionChains(driver)
            actions.click(element_to_click)
            actions.perform()

            time.sleep(1)

            html = driver.page_source
            html = BeautifulSoup(html,features="lxml")
            html = html.find('div', {'class':'logistics'})
            shipping = html.findAll('div', {'class':'table-tr'})
            ch = ['1','2','3','4','5','6','7','8','9','0',',']
            for i in shipping:
                sh = i
                if "AliExpress Standard Shipping" in str(i.get_text()):
                    sh = sh.findAll('div',{'class':'table-td'})
                    for k in sh:
                        if 'руб.' not in str(k):
                            sh.remove(k)
                        else:
                            k = str(k)
                            k = ''.join(j for j in k if j in ch)
                            k = k.replace(',','.')
                            k = float(k)
                            price = k

            price = str(price)
            price = float(price)
            percents = self.percentbox.text()
            percents = float(percents)
            price = price+price*(percents/100)
            price = format(price, '.2f')
            price = str(price)
            self.pricebox.setText(price)

        except Exception:
            self.pricebox.setText("Not found")

    
    @pyqtSlot()
    def standparse2(self):
        try:
            element_to_click = driver.find_element_by_xpath("//span[@class=\"product-shipping-info black-link\"]")
            actions = ActionChains(driver)
            actions.click(element_to_click)
            actions.perform()

            time.sleep(1)

            html = driver.page_source
            html = BeautifulSoup(html,features="lxml")
            html = html.find('div', {'class':'logistics'})
            shipping = html.findAll('div', {'class':'table-tr'})
            ch = ['1','2','3','4','5','6','7','8','9','0',',']
            for i in shipping:
                sh = i
                if "Cainiao Saver Shipping For Special Goods" in str(i.get_text()):
                    sh = sh.findAll('div',{'class':'table-td'})
                    for k in sh:
                        if 'руб.' not in str(k):
                            sh.remove(k)
                        else:
                            k = str(k)
                            k = ''.join(j for j in k if j in ch)
                            k = k.replace(',','.')
                            k = float(k)
                            price = k

            price = str(price)
            price = float(price)
            percents = self.percentbox.text()
            percents = float(percents)
            price = price+price*(percents/100)
            format(price, '.2f')
            price = str(price)
            self.pricebox.setText(price)

    
        except Exception:
            self.pricebox.setText("Not found")


    
    @pyqtSlot()
    def toexcelf(self):
        global imgs
        global keys
        wb = load_workbook(filename = 'products.xlsx')
        sheet_ranges = wb['Sheet1']
        i = 2
        while sheet_ranges[f'A{i}'].value!=None:
            i+=1
        sheet_ranges[f'A{i}'] = self.name.toPlainText()
        sheet_ranges[f'B{i}'] = self.description.toPlainText()
        sheet_ranges[f'C{i}'] = self.feature1.toPlainText()
        sheet_ranges[f'D{i}'] = self.feature2.toPlainText()
        sheet_ranges[f'E{i}'] = self.feature3.toPlainText()
        sheet_ranges[f'F{i}'] = self.feature4.toPlainText()
        sheet_ranges[f'G{i}'] = self.feature5.toPlainText()
        sheet_ranges[f'H{i}'] = self.pricebox.text()
        #sheet_ranges[f'I{i}'] = Image(imgs[0])
        #sheet_ranges[f'G{i}'] = Image(imgs[1])
        #sheet_ranges[f'K{i}'] = Image(imgs[2])
        #sheet_ranges[f'L{i}'] = Image(imgs[3])
        #sheet_ranges[f'M{i}'] = Image(imgs[4])
        sheet_ranges[f'N{i}'] = self.keys.text()


        wb.save("products.xlsx")
    
    @pyqtSlot()
    def deldubs(self):
        print('PyQt5 button click')
    
    @pyqtSlot()
    def getkeys(self):
        global keys
        chrome_options = Options()
        #chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        driver_helium = webdriver.Chrome(options=chrome_options)

        try:
            driver_helium.get("https://www.helium10.com/")
        except Exception:
            print('bruh')
        time.sleep(1)

        element_to_click = driver_helium.find_element_by_xpath("//a[@class=\"btn btn-outline-secondary btn-sm\"]")
        actions = ActionChains(driver_helium)
        actions.click(element_to_click)
        actions.perform()

        time.sleep(1)

        username = "vlad.mih.n92@gmail.com"
        password = "111Asq111"

        driver_helium.find_element_by_id("loginform-email").send_keys(username)
        driver_helium.find_element_by_id("loginform-password").send_keys(password)
        driver_helium.find_element_by_xpath("//button[@class=\"btn btn-secondary btn-block\"]").click()

        time.sleep(1)

        element = driver_helium.find_element_by_xpath("//nav[@class=\"left-nav\"]")
        actions = ActionChains(driver_helium)
        actions.move_to_element(element)
        actions.perform() 

        time.sleep(1)

        element_to_click = driver_helium.find_element_by_xpath("//div[@data-tool-id=\"keywordResearch\"]")
        actions = ActionChains(driver_helium)
        actions.click(element_to_click)
        actions.perform() 

        time.sleep(1)

        element_to_click = driver_helium.find_element_by_xpath("//a[@data-tool-id=\"keywordresearch\"]")
        actions = ActionChains(driver_helium)
        actions.click(element_to_click)
        actions.perform()   

        driver_helium.find_element_by_id("reversesearch-asin").send_keys(self.name.toPlainText())

        element_to_click = driver_helium.find_element_by_xpath("//a[@class=\"btn btn-primary reverse-search\"]")
        actions = ActionChains(driver_helium)
        actions.click(element_to_click)
        actions.perform()

        time.sleep(1)
        
        try:
            element_to_click = driver_helium.find_element_by_xpath("//div[@class=\"la-ball-fall\"]")
            actions = ActionChains(driver_helium)
            actions.click(element_to_click)
            actions.perform()  
        except Exception:
            print("bruh")

        time.sleep(8)

        element_to_click = driver_helium.find_element_by_xpath("//button[@class=\"btn btn-default dropdown-toggle tooltipstered\"]")
        actions = ActionChains(driver_helium)
        actions.click(element_to_click)
        actions.perform() 

        element_to_click = driver_helium.find_element_by_xpath("//div[@class=\"dropdown-item action-export-copy active\"]")
        actions = ActionChains(driver_helium)
        actions.click(element_to_click)
        actions.perform() 

        keys_s = pyperclip.paste()
        self.keys.insertPlainText(keys_s)
        self.show()

        
        
       

if __name__ == '__main__':
    p = 0
    app = QApplication(sys.argv)
    ex = App()
    sys.exit(app.exec_())